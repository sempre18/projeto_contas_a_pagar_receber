import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.sql import Session, Recebimento, Despesa


MONEY_FORMAT = '_-"R$ "* #,##0.00_-;-"R$ "* #,##0.00_-;_-"R$ "* \\-??_-;_-@_-'

CINZA_HEADER = "FFD9D9D9"       # header de datas do bloco DESPESAS
AZUL_HEADER = "FFBDD7EE"        # header de datas e linha TOTAL do bloco RECEBIMENTOS
AZUL_LINHA = "FFDDEBF6"         # linhas de categoria do bloco RECEBIMENTOS (cor única, sem ciclo)
VERDE_TOTAL = "FFE2EFDA"        # linha TOTAL do bloco DESPESAS
BRANCO = "FFFFFFFF"
PALETA_DESPESAS = ["FFE2EFDA", "FFFFE5FF", "FFDDEBF7", "FFFFF2CC", "FFE6CFB7"]

SEM_DESCRICAO = "(SEM DESCRIÇÃO)"
SEM_CONTA = "(SEM CONTA)"

# "Descrição Portador" tem centenas de variações de texto nos dados reais
# (espaços sobrando, notas de estorno/devolução etc.). Só bancos e maquininhas
# de cartão são agrupados; o resto vira sua própria categoria (normalizada).
PALAVRAS_BANCO = ["BANCO", "ITAU", "BRASIL", "BRADESCO", "SANTANDER", "CX ECONOM", "CAIXA ECON"]
PALAVRAS_CARTAO = ["REDE", "AZULZINHA"]


def _bucket_descricao_portador(valor):
    texto = (valor or "").strip().upper()
    if not texto:
        return SEM_DESCRICAO
    if any(palavra in texto for palavra in PALAVRAS_CARTAO):
        return "CARTÃO"
    if any(palavra in texto for palavra in PALAVRAS_BANCO):
        return "BOLETO"
    return texto


def _bucket_conta_centro_custo(valor):
    texto = (valor or "").strip().upper()
    return texto if texto else SEM_CONTA


def _dias_periodo(data_inicio, data_fim):
    dias = []
    dia = data_inicio
    while dia <= data_fim:
        dias.append(dia)
        dia += timedelta(days=1)
    return dias


def _somar_por_categoria_e_dia(model_class, coluna_categoria, coluna_valor, data_inicio, data_fim, bucket_fn):
    """{(categoria, dia): soma_do_valor} para as linhas de model_class com vencimento no período."""
    linhas = (
        Session.query(model_class.vencimento, coluna_categoria, coluna_valor)
        .filter(model_class.vencimento >= data_inicio, model_class.vencimento <= data_fim)
        .all()
    )
    totais = {}
    for dia, valor_categoria, valor in linhas:
        categoria = bucket_fn(valor_categoria)
        chave = (categoria, dia)
        totais[chave] = totais.get(chave, 0) + (valor or 0)
    return totais


def _estilizar(cell, *, bold=False, fill=None, align=None, border_top=None, border_bottom=None, number_format=None):
    cell.font = Font(bold=bold)
    if fill:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill)
    if align:
        cell.alignment = Alignment(horizontal=align)
    if border_top or border_bottom:
        cell.border = Border(
            top=Side(style=border_top) if border_top else None,
            bottom=Side(style=border_bottom) if border_bottom else None,
        )
    if number_format:
        cell.number_format = number_format


def _escrever_bloco(ws, linha, titulo, rotulo_coluna, totais, dias, cor_header, cores_linhas, cor_total):
    """
    Escreve um bloco (DESPESAS ou RECEBIMENTOS) a partir de `linha`.
    cores_linhas: uma cor única (str) pra todas as linhas, ou uma lista ciclada por linha.
    Retorna (linha_do_total, proxima_linha_livre).
    """
    n_dias = len(dias)
    col_total_periodo = n_dias + 2

    ultima_coluna = get_column_letter(col_total_periodo)
    ws.merge_cells(f"A{linha}:{ultima_coluna}{linha}")
    _estilizar(ws.cell(row=linha, column=1, value=titulo), bold=True, fill=BRANCO, align="center",
               border_top="medium", border_bottom="thin")
    for col in range(2, col_total_periodo + 1):
        _estilizar(ws.cell(row=linha, column=col), fill=BRANCO, border_top="medium", border_bottom="thin")
    linha += 1

    _estilizar(ws.cell(row=linha, column=1, value=rotulo_coluna), bold=True, fill=cor_header, align="left",
               border_top="thin", border_bottom="thin")
    for i, dia in enumerate(dias):
        _estilizar(ws.cell(row=linha, column=2 + i, value=dia.strftime("%d/%m/%Y")), bold=True, fill=cor_header,
                   align="center", border_top="thin", border_bottom="thin")
    _estilizar(ws.cell(row=linha, column=col_total_periodo, value="Total período"), bold=True, fill=cor_header,
               align="center", border_top="thin", border_bottom="thin")
    linha += 1

    categorias = sorted({categoria for categoria, _dia in totais})
    primeira_linha_categoria = linha
    col_b = get_column_letter(2)
    col_ultimo_dia = get_column_letter(1 + n_dias)

    for indice, categoria in enumerate(categorias):
        cor = cores_linhas if isinstance(cores_linhas, str) else cores_linhas[indice % len(cores_linhas)]
        _estilizar(ws.cell(row=linha, column=1, value=categoria), bold=True, fill=cor, align="left",
                   border_top="thin", border_bottom="thin")
        for i, dia in enumerate(dias):
            valor = totais.get((categoria, dia))
            _estilizar(ws.cell(row=linha, column=2 + i, value=valor), fill=cor, align="right",
                       border_top="thin", border_bottom="thin", number_format=MONEY_FORMAT)
        formula_total = f"=SUM({col_b}{linha}:{col_ultimo_dia}{linha})"
        _estilizar(ws.cell(row=linha, column=col_total_periodo, value=formula_total), fill=cor, align="right",
                   border_top="thin", border_bottom="thin", number_format=MONEY_FORMAT)
        linha += 1
    ultima_linha_categoria = linha - 1

    _estilizar(ws.cell(row=linha, column=1, value="TOTAL"), bold=True, fill=cor_total, align="left",
               border_top="medium", border_bottom="medium")
    for i in range(n_dias):
        col = 2 + i
        col_letra = get_column_letter(col)
        formula = f"=SUM({col_letra}{primeira_linha_categoria}:{col_letra}{ultima_linha_categoria})"
        _estilizar(ws.cell(row=linha, column=col, value=formula), bold=True, fill=cor_total, align="right",
                   border_top="medium", border_bottom="medium", number_format=MONEY_FORMAT)
    formula_total_geral = f"=SUM({col_b}{linha}:{col_ultimo_dia}{linha})"
    _estilizar(ws.cell(row=linha, column=col_total_periodo, value=formula_total_geral), bold=True, fill=cor_total,
               align="right", border_top="medium", border_bottom="medium", number_format=MONEY_FORMAT)

    linha_total = linha
    return linha_total, linha + 1


def _escrever_base(wb, nome_aba, model_class, data_inicio, data_fim):
    """Aba com os dados crus (banco de dados) que alimentam o bloco correspondente, filtrados pelo período."""
    colunas = [c for c in model_class.__table__.columns.keys() if c not in ("id", "_hash")]
    linhas = (
        Session.query(model_class)
        .filter(model_class.vencimento >= data_inicio, model_class.vencimento <= data_fim)
        .all()
    )
    ws = wb.create_sheet(nome_aba)
    ws.append(colunas)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for objeto in linhas:
        ws.append([getattr(objeto, coluna) for coluna in colunas])


def gerar_dashboard(caminho_saida, data_inicio, data_fim, incluir_bases=True):
    """
    Gera o dashboard "FLUXO DE CAIXA PREVISTO" em `caminho_saida`, com uma
    coluna por dia do período (data_inicio a data_fim, inclusive) mais uma
    coluna de total do período, e uma linha de total por bloco:

    - DESPESAS: agrupado por Conta Centro de Custo (Recebimento.conta_centro_custo),
      somando Vl.Documento (Recebimento.vl_documento).
    - RECEBIMENTOS: agrupado por Descrição Portador (Despesa.descrição_portador,
      bancos -> BOLETO, REDE/AZULZINHA -> CARTÃO), somando Valor Pago (Despesa.valor_pago).
    - SALDO PREVISTO FINAL DO DIA: Recebimentos - Despesas de cada dia.

    Se incluir_bases, adiciona abas com os dados crus de cada bloco no período.
    """
    dias = _dias_periodo(data_inicio, data_fim)
    n_dias = len(dias)
    col_total_periodo = n_dias + 2

    despesas_totais = _somar_por_categoria_e_dia(
        Recebimento, Recebimento.conta_centro_custo, Recebimento.vl_documento,
        data_inicio, data_fim, _bucket_conta_centro_custo,
    )
    recebimentos_totais = _somar_por_categoria_e_dia(
        Despesa, Despesa.descrição_portador, Despesa.valor_pago,
        data_inicio, data_fim, _bucket_descricao_portador,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 35
    for i in range(n_dias + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16.5

    ultima_coluna = get_column_letter(col_total_periodo)
    ws.merge_cells(f"A1:{ultima_coluna}1")
    _estilizar(ws.cell(row=1, column=1, value="FLUXO DE CAIXA PREVISTO"), bold=True, align="center")
    ws.merge_cells(f"A2:{ultima_coluna}2")
    texto_periodo = f"{data_inicio.strftime('%d/%m')} a {data_fim.strftime('%d/%m')}"
    _estilizar(ws.cell(row=2, column=1, value=texto_periodo), bold=True, align="center")

    linha = 5
    despesas_total_linha, linha = _escrever_bloco(
        ws, linha, "D E S P E S A S", "Conta Centro de Custo",
        despesas_totais, dias, CINZA_HEADER, PALETA_DESPESAS, VERDE_TOTAL,
    )
    linha += 2

    recebimentos_total_linha, linha = _escrever_bloco(
        ws, linha, "R E C E B I M E N T O S", "Descrição Portador",
        recebimentos_totais, dias, AZUL_HEADER, AZUL_LINHA, AZUL_HEADER,
    )
    linha += 2

    linha_saldo_header = linha
    linha_saldo_valor = linha + 1
    ws.merge_cells(f"A{linha_saldo_header}:A{linha_saldo_valor}")
    _estilizar(ws.cell(row=linha_saldo_header, column=1, value="SALDO PREVISTO FINAL DO DIA"), bold=True,
               fill=BRANCO, align="center", border_top="medium", border_bottom="thin")

    col_b = get_column_letter(2)
    col_ultimo_dia = get_column_letter(1 + n_dias)
    for i, dia in enumerate(dias):
        col = 2 + i
        col_letra = get_column_letter(col)
        _estilizar(ws.cell(row=linha_saldo_header, column=col, value=dia.strftime("%d/%m/%Y")), bold=True,
                   fill=BRANCO, align="center", border_top="medium", border_bottom="thin")
        formula = f"={col_letra}{recebimentos_total_linha}-{col_letra}{despesas_total_linha}"
        _estilizar(ws.cell(row=linha_saldo_valor, column=col, value=formula), fill=BRANCO, align="right",
                   border_top="thin", border_bottom="medium", number_format=MONEY_FORMAT)

    _estilizar(ws.cell(row=linha_saldo_header, column=col_total_periodo, value="Total período"), bold=True,
               fill=BRANCO, align="center", border_top="medium", border_bottom="thin")
    formula_total_saldo = f"=SUM({col_b}{linha_saldo_valor}:{col_ultimo_dia}{linha_saldo_valor})"
    _estilizar(ws.cell(row=linha_saldo_valor, column=col_total_periodo, value=formula_total_saldo), fill=BRANCO,
               align="right", border_top="thin", border_bottom="medium", number_format=MONEY_FORMAT)

    if incluir_bases:
        _escrever_base(wb, "Despesas - Base", Recebimento, data_inicio, data_fim)
        _escrever_base(wb, "Recebimentos - Base", Despesa, data_inicio, data_fim)

    wb.save(caminho_saida)


if __name__ == "__main__":
    pass
